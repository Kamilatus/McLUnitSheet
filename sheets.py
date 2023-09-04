import openpyxl as xl
from openpyxl.styles import Alignment, Side, Border
import mobs

columns_defense = ['name', 'health', 'speed', 'frequency', 'radius',
                   'general', 'mobs', 'buildables', 'walls']

columns_statuses = ['name', 'time', 'resistance time', 'damage', 'interval', 'damage types',
                    'speed mod', 'jump mod',
                    'disables movement and actions', 'negated statuses']


def clear_sheet(sheet):
    sheet.merge_cells(start_row=1, start_column=1, end_row=200, end_column=200)
    sheet.unmerge_cells(start_row=1, start_column=1, end_row=200, end_column=200)
    sheet.delete_cols(1, 200)


def create_sheet_attacks(sheet):
    clear_sheet(sheet)
    columns = ['mob name', 'attack name', 'damage', 'cooldown', 'min range', 'max range', 'damage types']
    for column in range(1, len(columns) + 1):
        sheet.cell(1, column).value = columns[column - 1]


def create_sheet_statuses(sheet):
    clear_sheet(sheet)
    sheet.cell(1, 4).value = 'damage over time'
    for column in range(1, len(columns_statuses) + 1):
        sheet.cell(2, column).value = columns_statuses[column - 1]
    sheet.merge_cells(start_row=1, start_column=4, end_row=1, end_column=6)


def create_sheet_defense(sheet):
    clear_sheet(sheet)
    for column in range(1, len(columns_defense) + 1):
        sheet.cell(2, column).value = columns_defense[column - 1]
    sheet.cell(1, 4).value = 'wander'
    sheet.cell(1, 6).value = 'targeting'
    sheet.cell(1, len(columns_defense) + 1).value = 'damage multipliers'


def create_sheets():
    workbook = xl.load_workbook('McL.xlsx')
    create_sheet_attacks(workbook['Friendly mobs - Attacks'])
    create_sheet_attacks(workbook['Piglins - Attacks'])
    create_sheet_defense(workbook['Friendly mobs - Defense'])
    create_sheet_defense(workbook['Piglins - Defense'])
    create_sheet_statuses(workbook['Statuses'])
    return workbook


def fill_sheet_attacks(mob_objects, sheet, game_path):
    columns = ['name', 'damage', 'cooldown', 'min range', 'max range']
    all_attacks = []
    row = 2
    for mob_object in mob_objects:
        sheet.cell(row, 1).value = mobs.name(mob_object)
        attacks = mobs.attacks(mob_object, game_path)
        all_attacks += attacks
        for attack in attacks:
            for column in range(2, len(columns) + 2):
                sheet.cell(row, column).value = attack[columns[column - 2]]
            if 'damage_types' in attack:
                for column2 in range(0, len(attack['damage_types'])):
                    sheet.cell(row, len(columns) + 2 + column2).value = attack['damage_types'][column2]
            row += 1
    notes_starting_column = sheet.max_column + 1
    row = 2
    for attack2 in all_attacks:
        fill_additional_notes(sheet, attack2['additional notes'], row, notes_starting_column, 1)
        row += 1


def fill_sheet_defense(mob_objects, sheet):
    unique_resistances = []
    row = 3
    max_column = len(columns_defense)
    for mob_object in mob_objects:
        mobs.defense(mob_object)
        for column in range(1, max_column + 1):
            sheet.cell(row, column).value = mob_object[columns_defense[column - 1]]
        for resistance in mob_object['resistances']:
            if resistance not in unique_resistances:
                unique_resistances.append(resistance)
                sheet.cell(2, max_column + len(unique_resistances)).value = resistance
            for column in range(0, len(unique_resistances)):
                if sheet.cell(2, max_column + column + 1).value == resistance:
                    sheet.cell(row, max_column + column + 1).value = mob_object['resistances'][resistance]
        row += 1


def fill_sheet_statuses(statuses, sheet):
    row = 3
    max_column = len(columns_statuses)
    for status in statuses:
        for column in range(1, max_column + 1):
            value = str(status[columns_statuses[column - 1]])
            if value != 'None':
                sheet.cell(row, column).value = value
            if 'additional notes' in status:
                fill_additional_notes(sheet, status['additional notes'], row, max_column + 1, 2)
        row += 1


def fill_additional_notes(sheet, notes, row, starting_column, title_row):
    sheet.cell(title_row, starting_column).value = 'additional notes'
    note_number = 0
    for note in notes:
        sheet.cell(row, starting_column + note_number).value = note
        note_number += 1


def format_sheets(workbook):
    format_regular_columns(workbook)
    format_notes_columns(workbook)
    format_borders(workbook)


def format_regular_columns(workbook):
    for sheet in workbook:
        max_column = sheet.max_column
        max_row = sheet.max_row
        for column in range(1, max_column + 1):
            sheet.column_dimensions[chr(96 + column)].width = column_width(sheet, column)
            sheet.column_dimensions[chr(96 + column)].width = column_width(sheet, column)
            for row in range(1, max_row + 1):
                sheet.cell(row, column).alignment = Alignment(horizontal='center')


def format_notes_columns(workbook):
    sheets_with_notes = [workbook['Friendly mobs - Attacks'], workbook['Piglins - Attacks'], workbook['Statuses']]
    for sheet in sheets_with_notes:
        # merge_title_cells(sheet)
        max_column = sheet.max_column
        max_row = sheet.max_row
        title_row = find_title_row(sheet)
        for column in range(1, max_column + 1):
            if sheet.cell(title_row, column).value == 'additional notes' or sheet.cell(title_row, column).value == 'negated statuses':
                for column2 in range(column, max_column + 1):
                    sheet.column_dimensions[chr(96 + column2)].width = 50
                for row in range(title_row + 1, max_row + 1):
                    for column2 in range(column + 1, max_column + 1):
                        col_width = cell_width(sheet.cell(row, column2).value)
                        next_col_width = cell_width(sheet.cell(row, column2 + 1).value)
                        if col_width > 50 and col_width > next_col_width > 0:
                            temp = sheet.cell(row, column2 + 1).value
                            sheet.cell(row, column2 + 1).value = sheet.cell(row, column).value
                            sheet.cell(row, column2).value = temp
                    for column2 in range(column, max_column + 1):
                        if cell_width(sheet.cell(row, column2).value) > 50:
                            if sheet.cell(row, column2 + 1).value is None:
                                sheet.cell(row, column2).alignment = Alignment(horizontal='left')
                            else:
                                sheet.cell(row, column2).alignment = Alignment(shrink_to_fit=True, horizontal='center')


def column_width(sheet, column):
    max_width = 0
    max_row = sheet.max_row
    for row in range(1, max_row):
        string = str(sheet.cell(row, column).value)
        width = cell_width(string)
        if width > max_width:
            max_width = width
    return max_width + 2


def cell_width(s):
    if s is None:
        return 0
    return len(s) - (s.count('\'') - s.count(',') - s.count('.') - s.count('[') - s.count(']') - s.count('{') - s.count('}')) * 0.3


def find_title_row(sheet):
    title_row = 2
    max_column = sheet.max_column
    for column in range(1, max_column + 1):
        if sheet.cell(1, column).value == 'additional notes':
            title_row = 1
    return title_row


def format_borders(workbook):
    for sheet in workbook:
        thin = Side(border_style='thin')
        double = Side(border_style='double')
        title_border = Border(left=thin, right=thin, bottom=double, top=thin)
        column_border = Border(left=thin, right=thin)
        bottom_border = Border(left=thin, right=thin, bottom=thin)
        mob_border = Border(left=thin, right=thin, top=thin)
        bottom_mob_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        title_row = find_title_row(sheet)
        if sheet.cell(2, 2).value == 'health':
            max_column = sheet.max_column  # idk why but borders of last column for defense sheets didn't work
        else:
            max_column = sheet.max_column - 1
        max_row = sheet.max_row
        for column in range(1, max_column + 1):
            sheet.cell(title_row, column).border = title_border
            sheet.cell(max_row, column).border = bottom_border
        if title_row == 2:
            for column in range(1, max_column + 1):
                if sheet.cell(1, column).value is not None:
                    sheet.cell(1, column).border = column_border
        for row in range(title_row + 1, max_row):
            for column in range(1, max_column + 1):
                sheet.cell(row, column).border = column_border
        if sheet.cell(1, 1).value == 'mob name':
            for row in range(title_row + 2, max_row + 1):
                if sheet.cell(row - 1, 1).value is None and sheet.cell(row, 1).value is not None:
                    for column in range(1, max_column + 1):
                        if row == max_row:
                            sheet.cell(row, column).border = bottom_mob_border
                        else:
                            sheet.cell(row, column).border = mob_border

# def merge_title_cells(sheet):
#     title_row = find_title_row(sheet)
#     max_column = sheet.max_column
#     cells_to_merge = {'additional notes': None, 'damage over time': None,
#                       'wander': None, 'targeting': None,
#                       'damage multipliers': None, 'damage types': None}
#     merge_area = {'damage over time': 2, 'wander': 1, 'targeting': 3, 'damage types': 0}
#     if 'Attacks' in sheet:
#         for column in range(1,max_column+1):
#             if sheet.cell(title_row,column).value == 'damage types':
#                 for column2 in range(column+1,max_column):
#                     if sheet.cell(title_row,column2).value == 'additional notes':
#                         merge_area['damage types'] = column2 - column
#     for row in range(1,title_row):
#         for column in range(1,max_column+1):
#             for cell in cells_to_merge:
#                 if sheet.cell(row,column).value == cell:
#                     cells_to_merge[cell] = sheet.cell
#     for cell in cells_to_merge:
#         cell = cells_to_merge[cell]
#         if cell is not None:
#             print(cell)
#             column = cell.column
#             row = cell.row
#             if cell.value in merge_area:
#                 column2 = column + merge_area[cell]
#             else:
#                 column2 = max_column
#             sheet.merge_cells(start_row=row,start_column=column,end_row=row,end_column=column2)
